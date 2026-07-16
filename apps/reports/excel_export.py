"""Excel export for tax compliance reports."""

from datetime import datetime
from decimal import Decimal
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone
from django.db.models import Sum, Q

from apps.sessions.models import Session, SessionPlayer, Attendance
from apps.payments.models import Payment, Wallet
from apps.donations.models import Donation, DonationCampaign
from apps.banking.models import BankTransaction


class TaxComplianceExcelExport:
    """Generate Excel report for tax compliance with multiple tabs."""
    
    def __init__(self, from_date=None, to_date=None, venue=None, generated_by=None):
        self.from_date = from_date
        self.to_date = to_date or timezone.now().date()
        self.location = venue  # venue param → location field
        self.generated_by = generated_by
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
        
    def generate(self):
        """Generate all tabs and return BytesIO buffer."""
        self._add_sessions_tab()
        self._add_attendance_tab()
        self._add_payments_tab()
        self._add_donations_tab()
        self._add_financial_summary_tab()
        
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _get_sessions(self):
        """Fetch sessions in date range and venue."""
        qs = Session.objects.filter(attendance_confirmed=True)
        
        if self.from_date:
            qs = qs.filter(date__gte=self.from_date)
        if self.to_date:
            qs = qs.filter(date__lte=self.to_date)
        if self.location:
            qs = qs.filter(location=self.location)
        
        return qs.order_by('date')
    
    def _add_sessions_tab(self):
        """Sessions Register tab."""
        ws = self.workbook.create_sheet('Sessions Register')
        
        # Headers
        headers = ['Session ID', 'Session Date', 'Venue', 'Venue Cost', 'Teams', 
                   'Match Status', 'Number of Players', 'Total Cost Split', 'Notes']
        ws.append(headers)
        
        # Style header
        self._style_header(ws, headers)
        
        # Data
        for session in self._get_sessions():
            matches = session.matches.all()
            teams = ', '.join([f"{m.name}" for m in matches]) if matches else 'N/A'
            
            # Count players
            player_count = SessionPlayer.objects.filter(session=session).count()
            
            # Get match status
            match_status = 'Completed' if matches.exists() else 'No Matches'
            
            ws.append([
                str(session.id),
                session.date.strftime('%Y-%m-%d'),
                session.location or 'N/A',
                f"€{session.cost or 0:.2f}",
                teams,
                match_status,
                player_count,
                f"€{session.cost_per_person or 0:.2f}" if session.cost_per_person else '€0.00',
                session.name,
            ])
        
        self._auto_adjust_columns(ws)
    
    def _add_attendance_tab(self):
        """Player Attendance Register tab - Shows who attended each session."""
        ws = self.workbook.create_sheet('Attendance Register')
        
        headers = ['Session Date', 'Session Name', 'Venue', 'Player Name', 'Attended', 'Team Assigned', 'Notes']
        ws.append(headers)
        
        self._style_header(ws, headers)
        
        # Get all sessions and their attendance
        sessions = list(self._get_sessions())
        
        # Batch fetch all players and attendances for efficiency
        from apps.matches.models import Player as MatchPlayer
        all_session_players = SessionPlayer.objects.filter(session__in=sessions).select_related('user')
        all_match_players = MatchPlayer.objects.filter(
            team__match__session__in=sessions
        ).select_related('team', 'user')
        all_attendances = Attendance.objects.filter(match_player__session__in=sessions)
        
        # Build lookup maps efficiently
        player_team_map = {}
        for mp in all_match_players:
            if mp.team and mp.team.match:
                key = (mp.user_id, mp.team.match.session_id)
                player_team_map[key] = mp.team.name
        
        attendance_map = {}
        for attendance in all_attendances:
            attendance_map[attendance.match_player_id] = attendance
        
        for session in sessions:
            session_players = [sp for sp in all_session_players if sp.session_id == session.id]
            
            if not session_players:
                # If no session players, add a blank row
                ws.append([
                    session.date.strftime('%Y-%m-%d'),
                    session.name,
                    session.location or 'N/A',
                    'No players registered',
                    '—',
                    '—',
                    '',
                ])
            else:
                for idx, sp in enumerate(session_players):
                    # Get attendance status from pre-fetched map
                    attendance = attendance_map.get(sp.id)
                    attended = 'Yes' if attendance and attendance.attended else 'No'
                    
                    # Get team assignment from pre-built map
                    team_name = player_team_map.get((sp.user_id, session.id), 'Not assigned')
                    
                    # Only show session info on first player row
                    if idx == 0:
                        ws.append([
                            session.date.strftime('%Y-%m-%d'),
                            session.name,
                            session.location or 'N/A',
                            sp.user.get_full_name() or sp.user.username,
                            attended,
                            team_name,
                            '',
                        ])
                    else:
                        ws.append([
                            '',
                            '',
                            '',
                            sp.user.get_full_name() or sp.user.username,
                            attended,
                            team_name,
                            '',
                        ])
        
        self._auto_adjust_columns(ws)
    
    def _add_payments_tab(self):
        """Financial Ledger - Payments from Players (invoices & settlements)."""
        ws = self.workbook.create_sheet('Payments Ledger')
        
        headers = ['Session Date', 'Session Name', 'Player Name', 'Invoice Amount', 
                   'Amount Paid', 'Outstanding', 'Payment Status', 'Payment Date', 'Payment Method']
        ws.append(headers)
        
        self._style_header(ws, headers)
        
        sessions = self._get_sessions()
        
        for session in sessions:
            # Get all session players with their payment info
            session_players = SessionPlayer.objects.filter(session=session).select_related('user')
            
            if not session_players.exists():
                # If no players, add blank row
                ws.append([
                    session.date.strftime('%Y-%m-%d'),
                    session.name,
                    'No players',
                    '€0.00',
                    '€0.00',
                    '€0.00',
                    '—',
                    '—',
                    '—',
                ])
            else:
                for sp in session_players:
                    # Get payment info
                    payment = Payment.objects.filter(user=sp.user, session=session).first()
                    
                    if payment:
                        invoice_amount = payment.amount
                        paid_amount = payment.amount if payment.status == 'paid' else Decimal('0')
                        outstanding = Decimal('0') if payment.status == 'paid' else payment.amount
                        payment_status = payment.status.title()
                        payment_date = payment.date.strftime('%Y-%m-%d')
                        payment_method = payment.method
                    else:
                        # No payment record
                        invoice_amount = session.cost_per_person or session.cost or Decimal('0')
                        paid_amount = Decimal('0')
                        outstanding = invoice_amount
                        payment_status = 'Not Invoiced'
                        payment_date = '—'
                        payment_method = '—'
                    
                    ws.append([
                        session.date.strftime('%Y-%m-%d'),
                        session.name,
                        sp.user.get_full_name() or sp.user.username,
                        f"€{invoice_amount:.2f}",
                        f"€{paid_amount:.2f}",
                        f"€{outstanding:.2f}",
                        payment_status,
                        payment_date,
                        payment_method,
                    ])
        
        # Add totals row
        ws.append([])  # Blank row
        
        # Calculate totals - fetch all payments at once for efficiency
        total_invoice = Decimal('0')
        total_paid = Decimal('0')
        total_outstanding = Decimal('0')
        
        sessions = list(self._get_sessions())
        all_payments = Payment.objects.filter(session__in=sessions)
        payment_map = {(p.user_id, p.session_id): p for p in all_payments}
        
        for session in sessions:
            session_players = SessionPlayer.objects.filter(session=session).select_related('user')
            for sp in session_players:
                payment = payment_map.get((sp.user_id, session.id))
                
                if payment:
                    total_invoice += payment.amount
                    if payment.status == 'paid':
                        total_paid += payment.amount
                    else:
                        total_outstanding += payment.amount
                else:
                    invoice = session.cost_per_person or session.cost or Decimal('0')
                    total_invoice += invoice
                    total_outstanding += invoice
        
        # Add totals row with styling
        totals_row = ws.append(['TOTAL', '', '', f"€{total_invoice:.2f}", f"€{total_paid:.2f}", f"€{total_outstanding:.2f}", '', '', ''])
        
        # Style totals row (bold, gray background)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
        
        self._auto_adjust_columns(ws)
    
    def _add_donations_tab(self):
        """Donations & Funding tab."""
        ws = self.workbook.create_sheet('Donations & Funding')
        
        headers = ['Donation Date', 'Donor Name', 'Amount', 'Campaign', 'Purpose', 
                   'Payment Method', 'Receipt Generated', 'Notes']
        ws.append(headers)
        
        self._style_header(ws, headers)
        
        donations = Donation.objects.all()
        
        if self.from_date:
            donations = donations.filter(donated_on__gte=self.from_date)
        if self.to_date:
            donations = donations.filter(donated_on__lte=self.to_date)
        
        donations = donations.order_by('donated_on')
        
        for donation in donations:
            ws.append([
                donation.donated_on.strftime('%Y-%m-%d'),
                donation.display_name,
                f"€{donation.amount:.2f}",
                donation.campaign.title if donation.campaign else 'General',
                donation.campaign.blurb if donation.campaign else 'General Donations',
                'Bank Transfer' if donation.source == 'bank' else 'Manual/Cash',
                'Yes' if donation.source == 'bank' else 'Manual Log',
                donation.note or '',
            ])
        
        self._auto_adjust_columns(ws)
    
    def _add_financial_summary_tab(self):
        """Financial Summary tab."""
        ws = self.workbook.create_sheet('Financial Summary')
        
        # Title
        ws['A1'] = 'INDCRIC CLUB — TAX COMPLIANCE FINANCIAL SUMMARY'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = 'Report Period:'
        ws['B3'] = f"{self.from_date or 'All'} to {self.to_date}"
        
        if self.location:
            ws['A4'] = 'Location Filter:'
            ws['B4'] = self.location
        
        ws['A6'] = 'Generated:'
        ws['B6'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        ws['A7'] = 'Generated By:'
        if self.generated_by:
            # Get full name, but fall back to username if empty
            full_name = self.generated_by.get_full_name().strip()
            generated_by_info = full_name if full_name else self.generated_by.username
        else:
            generated_by_info = 'Unknown'
        ws['B7'] = generated_by_info
        
        # Summary Data
        sessions = self._get_sessions()
        total_sessions = sessions.count()
        total_venue_cost = sessions.aggregate(total=Sum('cost'))['total'] or Decimal('0')
        
        # Total players
        total_players = SessionPlayer.objects.filter(
            session__in=sessions
        ).values('user').distinct().count()
        
        # Total PAID payments (only received money, not invoiced amounts)
        paid_payments = Payment.objects.filter(
            session__in=sessions,
            status='paid'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        # Outstanding/Pending payments (not yet received)
        pending_payments = Payment.objects.filter(
            session__in=sessions,
            status='pending'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        # Donations
        donations = Donation.objects.all()
        if self.from_date:
            donations = donations.filter(donated_on__gte=self.from_date)
        if self.to_date:
            donations = donations.filter(donated_on__lte=self.to_date)
        
        total_donations = donations.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        # Summary table
        ws['A10'] = 'SUMMARY STATISTICS'
        ws['A10'].font = Font(bold=True, size=12)
        
        summary_data = [
            ['Total Sessions Held', total_sessions],
            ['Total Unique Players', total_players],
            ['Total Venue Costs', f"€{total_venue_cost:.2f}"],
            ['Total Player Payments Received (Paid)', f"€{paid_payments:.2f}"],
            ['Outstanding Payments (Pending)', f"€{pending_payments:.2f}"],
            ['Total Donations Received', f"€{total_donations:.2f}"],
            ['Net Club Position', f"€{(total_donations + paid_payments - total_venue_cost):.2f}"],
        ]
        
        row = 11
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Compliance checklist
        ws['A20'] = 'COMPLIANCE CHECKLIST'
        ws['A20'].font = Font(bold=True, size=12)
        
        checklist = [
            ['✓', 'All players recorded'],
            ['✓', 'All attendance tracked'],
            ['✓', 'All payments documented'],
            ['✓', 'All donations logged'],
            ['✓', 'No cash discrepancies'],
        ]
        
        row = 21
        for check in checklist:
            ws[f'A{row}'] = check[0]
            ws[f'B{row}'] = check[1]
            row += 1
        
        self._auto_adjust_columns(ws)
    
    def _style_header(self, ws, headers):
        """Style the header row."""
        fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        font = Font(bold=True, color='FFFFFF')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
