from functools import wraps
from io import BytesIO
from collections import defaultdict

from django.conf import settings
from django.core.exceptions import ValidationError
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.http import url_has_allowed_host_and_scheme
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .forms import JerseyOrderForm, JerseyOrderWindowForm
from .models import JerseyOrder, JerseyOrderWindow


def jersey_ordering_enabled(view):
    """Gate a jersey view behind the JERSEY_ORDERING_ENABLED flag. When off
    (e.g. prod until prices are final), redirect home instead of serving it."""
    @wraps(view)
    def wrapper(request, *args, **kwargs):
        if not settings.JERSEY_ORDERING_ENABLED:
            messages.info(request, "Jersey ordering isn't available yet.")
            return redirect('home')
        return view(request, *args, **kwargs)
    return wrapper


def _catalog():
    catalog = []
    for code, label in JerseyOrder.ITEM_CHOICES:
        meta = JerseyOrder.ITEM_META.get(code, {})
        catalog.append({
            'code': code,
            'label': label,
            'rate': JerseyOrder.rate_for(code),
            'is_shirt': code in JerseyOrder.SHIRT_ITEMS,
            'visual': meta.get('visual', 'Item'),
            'group': meta.get('group', 'Kit'),
            'note': meta.get('note', ''),
        })
    return catalog


def _taken_numbers():
    references = {}
    orders = (
        JerseyOrder.objects
        .filter(for_person=JerseyOrder.FOR_SELF)
        .exclude(jersey_number='')
        .select_related('user')
        .order_by('jersey_number', 'wearer_name', 'user__username')
    )
    for order in orders:
        key = (order.jersey_number,
               order.wearer_name.strip().lower(), order.user_id)
        if key not in references:
            references[key] = {
                'jersey_number': order.jersey_number,
                'wearer_name': order.wearer_name,
                'gender': order.get_gender_display(),
                'user__username': order.user.username,
                'item_count': 0,
            }
        references[key]['item_count'] += 1
    rows = list(references.values())
    rows.sort(key=lambda r: (len(r['jersey_number']), r['jersey_number']))
    return rows


def _summary():
    rows = (
        JerseyOrder.objects
        .values('item_type', 'size', 'gender')
        .annotate(quantity=Sum('quantity'))
        .order_by('item_type', 'gender', 'size')
    )
    item_labels = dict(JerseyOrder.ITEM_CHOICES)
    gender_labels = dict(JerseyOrder.GENDER_CHOICES)
    grouped = defaultdict(list)
    grand_total = 0
    total_quantity = 0
    for row in rows:
        rate = JerseyOrder.rate_for(row['item_type'])
        quantity = row['quantity'] or 0
        total = rate * quantity
        grand_total += total
        total_quantity += quantity
        grouped[row['item_type']].append({
            'item': item_labels.get(row['item_type'], row['item_type']),
            'size': 'Free size - cap/hat' if row['size'] == 'FS' else row['size'],
            'gender': gender_labels.get(row['gender'], row['gender']),
            'quantity': quantity,
            'rate': rate,
            'total': total,
        })
    return grouped, total_quantity, grand_total


@login_required
@jersey_ordering_enabled
def jersey_orders_view(request):
    form = JerseyOrderForm(user=request.user)
    order_window = JerseyOrderWindow.current()
    ordering_open = JerseyOrderWindow.ordering_is_open()
    ordering_status = (
        order_window.status_text()
        if order_window
        else 'Ordering is open. Admin can set an order cutoff from Django admin.'
    )
    ordering_deadline = (
        order_window.closes_at_label()
        if order_window and order_window.is_enabled and order_window.closes_at
        else ''
    )
    if request.method == 'POST':
        if not ordering_open:
            messages.error(
                request, 'Jersey ordering is closed. No new orders can be added now.')
            return redirect('jersey-orders')
        form = JerseyOrderForm(request.POST, user=request.user)
        if form.is_valid():
            orders = form.save_orders()
            messages.success(
                request,
                f"Added {len(orders)} item{'' if len(orders) == 1 else 's'} for {form.cleaned_data['wearer_name']}."
            )
            return redirect('jersey-orders')

    own_orders = list(JerseyOrder.objects.filter(
        user=request.user).order_by('-created_at'))
    own_order_total = sum((order.line_total for order in own_orders), start=0)
    own_order_quantity = sum((order.quantity for order in own_orders), start=0)
    # Size choices per own-order line, for the inline editor (adult shirt/pant only).
    for o in own_orders:
        if o.item_type in JerseyOrder.SHIRT_ITEMS:
            o.size_choices = JerseyOrder.SHIRT_SIZE_CHOICES
        elif o.item_type in JerseyOrder.PANT_ITEMS:
            o.size_choices = JerseyOrder.PANT_SIZE_CHOICES
        else:
            o.size_choices = None
    context = {
        'form': form,
        'own_orders': own_orders,
        'own_order_total': own_order_total,
        'own_order_quantity': own_order_quantity,
        'catalog': _catalog(),
        'item_rows': form.item_rows(),
        'shirt_size_measurements': JerseyOrder.SIZE_MEASUREMENTS,
        'pant_size_measurements': JerseyOrder.PANT_SIZE_MEASUREMENTS,
        'taken_numbers': _taken_numbers(),
        'ordering_open': ordering_open,
        'ordering_status': ordering_status,
        'ordering_deadline': ordering_deadline,
    }
    return render(request, 'jerseys/order_form.html', context)


@login_required
@jersey_ordering_enabled
def edit_jersey_order_view(request, order_id):
    order = JerseyOrder.objects.filter(id=order_id).first()
    if order is None:
        messages.info(request, 'That order was already removed.')
        return redirect('jersey-orders')
    if not request.user.is_staff and order.user_id != request.user.id:
        messages.error(request, "You cannot edit another member's order.")
        return redirect('jersey-orders')
    if not request.user.is_staff and not JerseyOrderWindow.ordering_is_open():
        messages.error(
            request, 'Jersey ordering is closed. Existing orders can no longer be changed.')
        return redirect('jersey-orders')
    if request.method == 'POST':
        try:
            qty = int(request.POST.get('quantity') or 0)
        except (TypeError, ValueError):
            qty = 0
        if qty < 1:
            messages.error(request, 'Quantity must be at least 1.')
            return redirect('jersey-orders')
        order.quantity = qty
        new_size = (request.POST.get('size') or '').strip()
        if new_size and order.item_type in (JerseyOrder.SHIRT_ITEMS | JerseyOrder.PANT_ITEMS):
            order.size = new_size
        try:
            order.full_clean()
            order.save()
            messages.success(request, 'Order updated.')
        except ValidationError as exc:
            msgs = '; '.join(' '.join(v) for v in exc.message_dict.values())
            messages.error(request, msgs or 'Could not update the order.')
    return redirect('jersey-orders')


@login_required
@jersey_ordering_enabled
def delete_jersey_order_view(request, order_id):
    next_url = request.POST.get('next') or request.GET.get('next')

    def _dest():
        if next_url and url_has_allowed_host_and_scheme(
            next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure(),
        ):
            return redirect(next_url)
        return redirect('jersey-orders-admin' if request.user.is_staff else 'jersey-orders')

    order = JerseyOrder.objects.filter(id=order_id).first()
    if order is None:
        # Stale page (already removed) — don't 404, just go back with a note.
        messages.info(request, 'That order was already removed.')
        return _dest()
    if not request.user.is_staff and order.user_id != request.user.id:
        messages.error(request, "You cannot delete another member's order.")
        return redirect('jersey-orders')
    if not request.user.is_staff and not JerseyOrderWindow.ordering_is_open():
        messages.error(
            request, 'Jersey ordering is closed. Existing orders can no longer be changed.')
        return redirect('jersey-orders')
    if request.method == 'POST':
        owner, wearer = order.user, order.wearer_name
        order.delete()
        # Recompute the wearer's reference (e.g. if the numbered order was removed).
        JerseyOrder.sync_reference(owner, wearer)
        messages.success(request, 'Order line removed.')
    return _dest()


@staff_member_required
@jersey_ordering_enabled
def jersey_orders_admin_view(request):
    summary, total_quantity, grand_total = _summary()
    order_window = JerseyOrderWindow.current()
    window_form = JerseyOrderWindowForm(instance=order_window)
    if request.method == 'POST' and request.POST.get('action') == 'update_order_window':
        window_form = JerseyOrderWindowForm(
            request.POST, instance=order_window)
        if window_form.is_valid():
            order_window = window_form.save()
            messages.success(request, 'Jersey order close date updated.')
            return redirect('jersey-orders-admin')

    orders = JerseyOrder.objects.select_related('user').order_by(
        'user__first_name',
        'user__username',
        'wearer_name',
    )
    # Flat rows for the client-side sortable/filterable member-wise table.
    orders_data = [{
        'member': (o.user.get_full_name() or o.user.username) if o.user else '—',
        'forp': o.get_for_person_display(),
        'gender': o.get_gender_display(),
        'wearer': o.wearer_name,
        'item': o.get_item_type_display(),
        'size': o.display_size,
        'qty': o.quantity,
        'ref': o.reference or '',
        'number': o.jersey_number or '',
        'total': float(o.line_total),
    } for o in orders]
    return render(request, 'jerseys/admin_summary.html', {
        'orders': orders,
        'orders_data': orders_data,
        'summary': dict(summary),
        'total_quantity': total_quantity,
        'grand_total': grand_total,
        'taken_numbers': _taken_numbers(),
        'ordering_open': JerseyOrderWindow.ordering_is_open(),
        'ordering_status': order_window.status_text() if order_window else 'Ordering is open.',
        'ordering_deadline': (
            order_window.closes_at_label()
            if order_window and order_window.is_enabled and order_window.closes_at
            else ''
        ),
        'order_window_form': window_form,
    })


@staff_member_required
@jersey_ordering_enabled
def export_jersey_orders_view(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Jersey Orders'
    headers = [
        'Member',
        'For',
        'Gender',
        'Wearer name',
        'Item',
        'Size / measurement',
        'Shirt Size',
        'Pant Size',
        'Kid Measurements',
        'Quantity',
        'Jersey number',
        'Unit price',
        'Line total',
        'Notes',
    ]
    worksheet.append(headers)
    header_fill = PatternFill(fill_type='solid', fgColor='D9EAD3')
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for order in JerseyOrder.objects.select_related('user').order_by('user__username', 'wearer_name'):
        worksheet.append([
            order.user.get_full_name() or order.user.username,
            order.get_for_person_display(),
            order.get_gender_display(),
            order.wearer_name,
            order.get_item_type_display(),
            order.display_size,
            order.display_size if order.item_type in JerseyOrder.SHIRT_ITEMS and not order.display_size.startswith(
                'Kid custom') else '',
            order.display_size if order.item_type in JerseyOrder.PANT_ITEMS and not order.display_size.startswith(
                'Kid custom') else '',
            order.display_size if order.display_size.startswith(
                'Kid custom') else '',
            order.quantity,
            order.jersey_number,
            order.unit_price,
            order.line_total,
            order.notes,
        ])

    for column_cells in worksheet.columns:
        width = max(len(str(cell.value or '')) for cell in column_cells) + 2
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(
            width, 40)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="jersey_orders.xlsx"'
    return response
