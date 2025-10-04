from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q, Count
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from decimal import Decimal
from datetime import datetime, date
import json

from .models import (
    User, Match, MonthlyPeriod, PlayerBalance,
    EnhancedAttendance
)


@login_required
def enhanced_attendance_view(request):
    """Enhanced attendance view similar to R application"""

    # Get or create current monthly period
    current_date = timezone.now().date()
    current_month_name = current_date.strftime("%B %Y")

    try:
        current_period = MonthlyPeriod.objects.get(name=current_month_name)
    except MonthlyPeriod.DoesNotExist:
        # Create current month period
        current_period = MonthlyPeriod.objects.create(
            name=current_month_name,
            start_date=current_date.replace(day=1),
            # Simplified, should handle month end properly
            end_date=current_date.replace(day=28)
        )

    # Get all periods for dropdown
    periods = MonthlyPeriod.objects.all()

    # Get selected period from request
    selected_period_id = request.GET.get('period_id', current_period.id)
    try:
        selected_period = MonthlyPeriod.objects.get(id=selected_period_id)
    except MonthlyPeriod.DoesNotExist:
        selected_period = current_period

    # Get all active users
    users = User.objects.filter(is_active=True).order_by('username')

    # Get matches in the selected period
    matches = Match.objects.filter(
        date__gte=selected_period.start_date,
        date__lte=selected_period.end_date
    ).order_by('date')

    # Get or create player balances for the selected period
    player_balances = {}
    for user in users:
        balance, created = PlayerBalance.objects.get_or_create(
            user=user,
            period=selected_period,
            defaults={
                'previous_balance': Decimal('0.00'),
                'monthly_advance': Decimal('0.00'),
                'total_advance': Decimal('0.00'),
                'balance': Decimal('0.00')
            }
        )
        player_balances[user.id] = balance

    # Get attendance data for the period
    attendance_data = {}
    for user in users:
        attendance_data[user.id] = {}
        for match in matches:
            try:
                attendance = EnhancedAttendance.objects.get(
                    user=user, match=match)
                attendance_data[user.id][match.id] = attendance.status
            except EnhancedAttendance.DoesNotExist:
                attendance_data[user.id][match.id] = 'E'  # Default to Excused

    # Calculate totals
    totals = calculate_period_totals(selected_period, users, matches)

    # Handle POST requests
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_attendance':
            return handle_attendance_update(request, selected_period)
        elif action == 'set_advance':
            return handle_advance_update(request, selected_period)
        elif action == 'add_match':
            return handle_add_match(request, selected_period)
        elif action == 'create_period':
            return handle_create_period(request)

    context = {
        'periods': periods,
        'selected_period': selected_period,
        'users': users,
        'matches': matches,
        'player_balances': player_balances,
        'attendance_data': attendance_data,
        'totals': totals,
        'current_date': current_date,
    }

    return render(request, 'cric/pages/enhanced_attendance.html', context)


def calculate_period_totals(period, users, matches):
    """Calculate totals for the period"""
    totals = {
        'total_previous_balance': Decimal('0.00'),
        'total_monthly_advance': Decimal('0.00'),
        'total_advance': Decimal('0.00'),
        'total_balance': Decimal('0.00'),
        'match_totals': {},
        'match_counts': {}
    }

    # Calculate user totals
    for user in users:
        try:
            balance = PlayerBalance.objects.get(user=user, period=period)
            totals['total_previous_balance'] += balance.previous_balance
            totals['total_monthly_advance'] += balance.monthly_advance
            totals['total_advance'] += balance.total_advance
            totals['total_balance'] += balance.balance
        except PlayerBalance.DoesNotExist:
            pass

    # Calculate match totals
    for match in matches:
        # Total cost for the match
        present_count = EnhancedAttendance.objects.filter(
            match=match, status='P'
        ).count()

        cost_per_person = match.cost_per_person or Decimal('0.00')
        total_cost = present_count * cost_per_person

        totals['match_totals'][match.id] = total_cost
        totals['match_counts'][match.id] = present_count

    return totals


def handle_attendance_update(request, period):
    """Handle attendance updates"""
    try:
        match_id = request.POST.get('match_id')
        match = get_object_or_404(Match, id=match_id)

        # Get selected users and attendance status
        selected_users = request.POST.getlist('selected_users')
        attendance_status = request.POST.get('attendance_status', 'P')
        update_mode = request.POST.get('update_mode', 'selected_only')

        if update_mode == 'selected_only':
            # Update only selected users
            for user_id in selected_users:
                user = get_object_or_404(User, id=user_id)
                attendance, created = EnhancedAttendance.objects.get_or_create(
                    user=user, match=match,
                    defaults={'status': attendance_status}
                )
                if not created:
                    attendance.status = attendance_status
                    attendance.save()

        elif update_mode == 'selected_present_rest_excused':
            # Set selected users as Present, others as Excused
            all_users = User.objects.filter(is_active=True)
            for user in all_users:
                status = 'P' if str(user.id) in selected_users else 'E'
                attendance, created = EnhancedAttendance.objects.get_or_create(
                    user=user, match=match,
                    defaults={'status': status}
                )
                if not created:
                    attendance.status = status
                    attendance.save()

        # Recalculate all balances for the period
        recalculate_period_balances(period)

        messages.success(request, 'Attendance updated successfully!')

    except Exception as e:
        messages.error(request, f'Error updating attendance: {str(e)}')

    return redirect('enhanced_attendance')


def handle_advance_update(request, period):
    """Handle monthly advance updates"""
    try:
        selected_users = request.POST.getlist('selected_users')
        advance_amount = Decimal(request.POST.get('advance_amount', '0.00'))

        for user_id in selected_users:
            user = get_object_or_404(User, id=user_id)
            balance, created = PlayerBalance.objects.get_or_create(
                user=user, period=period
            )
            balance.monthly_advance += advance_amount
            balance.calculate_balance()

        messages.success(
            request, f'Added €{advance_amount} advance to {len(selected_users)} player(s)!')

    except Exception as e:
        messages.error(request, f'Error updating advances: {str(e)}')

    return redirect('enhanced_attendance')


def handle_add_match(request, period):
    """Handle adding a new match"""
    try:
        match_date = datetime.strptime(
            request.POST.get('match_date'), '%Y-%m-%d').date()
        cost_per_person = Decimal(request.POST.get('cost_per_person', '4.50'))
        match_name = request.POST.get(
            'match_name', f'Match {match_date.strftime("%d/%m/%Y")}')

        # Check if date is within period
        if not (period.start_date <= match_date <= period.end_date):
            messages.error(
                request, 'Match date must be within the selected period!')
            return redirect('enhanced_attendance')

        match = Match.objects.create(
            name=match_name,
            date=match_date,
            time=timezone.now().time(),  # Default time
            cost=cost_per_person * 20,  # Estimate for 20 players
            cost_per_person=cost_per_person,
            duration=3,  # Default 3 hours
            location='TBD'
        )

        messages.success(request, f'Match "{match_name}" added successfully!')

    except Exception as e:
        messages.error(request, f'Error adding match: {str(e)}')

    return redirect('enhanced_attendance')


def handle_create_period(request):
    """Handle creating a new monthly period"""
    try:
        period_name = request.POST.get('period_name')
        start_date = datetime.strptime(
            request.POST.get('start_date'), '%Y-%m-%d').date()
        end_date = datetime.strptime(
            request.POST.get('end_date'), '%Y-%m-%d').date()

        # Check if period already exists
        if MonthlyPeriod.objects.filter(name=period_name).exists():
            messages.error(request, 'A period with this name already exists!')
            return redirect('enhanced_attendance')

        # Create new period
        new_period = MonthlyPeriod.objects.create(
            name=period_name,
            start_date=start_date,
            end_date=end_date
        )

        # Copy balances from previous period
        try:
            previous_period = MonthlyPeriod.objects.filter(
                end_date__lt=start_date
            ).order_by('-end_date').first()

            if previous_period:
                copy_balances_to_new_period(previous_period, new_period)
        except Exception as e:
            print(
                f"Warning: Could not copy balances from previous period: {e}")

        messages.success(
            request, f'New period "{period_name}" created successfully!')

    except Exception as e:
        messages.error(request, f'Error creating period: {str(e)}')

    return redirect('enhanced_attendance')


def copy_balances_to_new_period(old_period, new_period):
    """Copy balances from old period to new period"""
    old_balances = PlayerBalance.objects.filter(period=old_period)

    for old_balance in old_balances:
        PlayerBalance.objects.create(
            user=old_balance.user,
            period=new_period,
            # Previous balance becomes the balance from last period
            previous_balance=old_balance.balance,
            monthly_advance=Decimal('0.00'),
            total_advance=old_balance.balance,
            balance=old_balance.balance
        )


def recalculate_period_balances(period):
    """Recalculate all player balances for a period"""
    balances = PlayerBalance.objects.filter(period=period)
    for balance in balances:
        balance.calculate_balance()


@login_required
def export_period_excel(request, period_id):
    """Export period data to Excel (similar to R app download)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from django.http import HttpResponse

        period = get_object_or_404(MonthlyPeriod, id=period_id)
        users = User.objects.filter(is_active=True).order_by('username')
        matches = Match.objects.filter(
            date__gte=period.start_date,
            date__lte=period.end_date
        ).order_by('date')

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{period.name} Attendance"

        # Headers
        headers = ['Name', 'PrevBalance', 'MonthAdvance', 'TotalAdvance']
        for match in matches:
            headers.append(
                f"{match.date.strftime('%d/%m/%Y')} (€{match.cost_per_person})")
        headers.append('Balance')

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write data
        row = 2
        totals_row = [0] * len(headers)
        totals_row[0] = "Total"

        for user in users:
            col = 1
            ws.cell(row=row, column=col, value=user.username)
            col += 1

            # Get balance data
            try:
                balance = PlayerBalance.objects.get(user=user, period=period)
                ws.cell(row=row, column=col, value=float(
                    balance.previous_balance))
                totals_row[1] += float(balance.previous_balance)
                col += 1

                ws.cell(row=row, column=col, value=float(
                    balance.monthly_advance))
                totals_row[2] += float(balance.monthly_advance)
                col += 1

                ws.cell(row=row, column=col, value=float(
                    balance.total_advance))
                totals_row[3] += float(balance.total_advance)
                col += 1

            except PlayerBalance.DoesNotExist:
                ws.cell(row=row, column=col, value=0.00)
                col += 1
                ws.cell(row=row, column=col, value=0.00)
                col += 1
                ws.cell(row=row, column=col, value=0.00)
                col += 1

            # Attendance data
            for i, match in enumerate(matches):
                try:
                    attendance = EnhancedAttendance.objects.get(
                        user=user, match=match)
                    ws.cell(row=row, column=col, value=attendance.status)
                    if attendance.status == 'P':
                        totals_row[4 + i] += float(match.cost_per_person or 0)
                except EnhancedAttendance.DoesNotExist:
                    ws.cell(row=row, column=col, value='E')
                col += 1

            # Balance
            try:
                balance = PlayerBalance.objects.get(user=user, period=period)
                ws.cell(row=row, column=col, value=float(balance.balance))
                totals_row[-1] += float(balance.balance)
            except PlayerBalance.DoesNotExist:
                ws.cell(row=row, column=col, value=0.00)

            row += 1

        # Write totals row
        for col, total in enumerate(totals_row, 1):
            cell = ws.cell(row=row, column=col, value=total)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="FFFF99", end_color="FFFF99", fill_type="solid")

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{period.name}_attendance.xlsx"'

        wb.save(response)
        return response

    except ImportError:
        messages.error(
            request, 'openpyxl package is required for Excel export')
        return redirect('enhanced_attendance')
    except Exception as e:
        messages.error(request, f'Error exporting data: {str(e)}')
        return redirect('enhanced_attendance')


@login_required
def ajax_update_attendance(request):
    """AJAX endpoint for quick attendance updates"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            user_id = data.get('user_id')
            match_id = data.get('match_id')
            status = data.get('status')

            user = get_object_or_404(User, id=user_id)
            match = get_object_or_404(Match, id=match_id)

            attendance, created = EnhancedAttendance.objects.get_or_create(
                user=user, match=match,
                defaults={'status': status}
            )

            if not created:
                attendance.status = status
                attendance.save()

            # Recalculate balance
            period = MonthlyPeriod.objects.filter(
                start_date__lte=match.date,
                end_date__gte=match.date
            ).first()

            if period:
                try:
                    balance = PlayerBalance.objects.get(
                        user=user, period=period)
                    new_balance = balance.calculate_balance()
                    return JsonResponse({
                        'success': True,
                        'new_balance': str(new_balance)
                    })
                except PlayerBalance.DoesNotExist:
                    pass

            return JsonResponse({'success': True})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})
